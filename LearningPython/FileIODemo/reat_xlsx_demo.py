from openpyxl import Workbook,load_workbook

wb=load_workbook(filename='demoexcel.xlsx')
#sh=wb.active
sh=wb['Sheet']#if we want to specify the sheet name specifically this is how we can
# print(sh['A3'].value)
# print(wb['Sheet']['A2'].value)
# print(sh.cell(2,2).value)
# print(sh.cell(row=3,column=2).value)
row_count=sh.max_row
col_count=sh.max_column
print(row_count)
print(col_count)

for i in range(1, row_count+1):
    for j in range(1, col_count+1):
        print(sh.cell(row=i,column=j).value, end=' ')
    print('\n')