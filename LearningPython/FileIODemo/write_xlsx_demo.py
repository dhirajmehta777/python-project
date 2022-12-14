from openpyxl import Workbook
wb  = Workbook()
ws = wb.active
#ws['A1']="python learning"
############################################################
# testdata=[['Name','City'],['Manish','Bangalore'],['Rayan','Maxico'],['Dhiraj','Mumbai']]
# for data in testdata:
#     ws.append(data)
##############################################################
for i in range(1, 6):
    for j in range(1, 4):
        ws.cell(row=i,column=j).value=i+j
wb.save("demoexcel.xlsx")

