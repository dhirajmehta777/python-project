from faker import Faker
from openpyxl import Workbook
wb=Workbook()
ws=wb.active
#fake_data=Faker()
#fake_data=Faker(['hi_IN','en_US'])#if we want to generate in specific language
fake_data=Faker('en_US')#if we want to generate in english language
for i in range(1, 11):
    for j in range(1, 4):
        ws.cell(row=i,column=1).value=fake_data.name()
        ws.cell(row=i, column=2).value = fake_data.email()
        ws.cell(row=i, column=3).value = fake_data.address()
wb.save("testdata.xlsx")
    #print(fake_data.name())
# print(fake_data.name())
# print(fake_data.email())
# print(fake_data.address())